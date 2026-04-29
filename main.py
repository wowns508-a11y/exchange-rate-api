from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
import requests
import re
from datetime import datetime, timedelta
import urllib3
import os
import bcrypt
import time
from supabase import create_client
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl import Workbook
from io import BytesIO

urllib3.disable_warnings()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

SMBS_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "http://www.smbs.biz/ExRate/StdExRate.jsp"
}

ALL_TARGET = ["USD", "CNH", "KWD", "AED", "SAR", "KZT", "MXN"]

CUR_NAMES = {
    "USD": "미국 달러", "CNH": "중국 위안화",
    "KWD": "쿠웨이트 디나르", "AED": "UAE 디르함",
    "SAR": "사우디 리얄", "KZT": "카자흐스탄 텡게",
    "MXN": "멕시코 페소", "IQD": "이라크 디나르",
    "LBP": "레바논 파운드",
}

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# =====================
# 캐시
# =====================
_pnl_cache:   dict = {}
_rates_cache: dict = {}
PNL_CACHE_TTL   = 1800  # 30분
RATES_CACHE_TTL = 1800  # 30분

def get_cached_records() -> list:
    now = time.time()
    if "records" in _pnl_cache and now - _pnl_cache["timestamp"] < PNL_CACHE_TTL:
        print(f"[PNL CACHE HIT] {len(_pnl_cache['records'])}건")
        return _pnl_cache["records"]
    print("[PNL CACHE MISS] Supabase 재조회")
    records = fetch_supabase_all()
    _pnl_cache["records"]   = records
    _pnl_cache["timestamp"] = now
    print(f"[PNL CACHE SET] {len(records)}건")
    return records

def get_cached_rates() -> dict | None:
    now = time.time()
    if "data" in _rates_cache and now - _rates_cache["timestamp"] < RATES_CACHE_TTL:
        print("[RATES CACHE HIT]")
        return _rates_cache["data"]
    return None

def set_rates_cache(data: dict):
    _rates_cache["data"]      = data
    _rates_cache["timestamp"] = time.time()
    print("[RATES CACHE SET]")

# =====================
# Supabase PnL 조회 (Airtable 대체)
# =====================
def fetch_supabase_all() -> list:
    """
    pnl_monthly + branches 조인해서 Framer가 기대하는 한국어 필드명으로 변환
    """
    all_records = []
    page_size   = 1000
    offset      = 0

    while True:
        res = supabase.table("pnl_monthly").select(
            "year, month, revenue, material_cost, labor_cost, "
            "expenses, hq_allocated_cost, gross_profit, operating_profit, "
            "ops_labor_cost, ops_expenses, division_cost, operating_profit_2, "
            "branches(branch_name, entity_name)"
        ).range(offset, offset + page_size - 1).execute()

        rows = res.data
        if not rows:
            break

        for row in rows:
            branch  = row.get("branches") or {}
            revenue = float(row.get("revenue", 0) or 0)
            mat     = float(row.get("material_cost", 0) or 0)
            labor   = float(row.get("labor_cost", 0) or 0)
            exp     = float(row.get("expenses", 0) or 0)
            corp    = float(row.get("hq_allocated_cost", 0) or 0)
            gross   = float(row.get("gross_profit", 0) or 0)
            op      = float(row.get("operating_profit", 0) or 0)
            ops_labor = float(row.get("ops_labor_cost", 0) or 0)
            ops_exp   = float(row.get("ops_expenses", 0) or 0)
            div_cost  = float(row.get("division_cost", 0) or 0)
            op2       = float(row.get("operating_profit_2", 0) or 0)
            m       = revenue if revenue != 0 else 1  # 분모 0 방지

            year  = int(row.get("year", 0) or 0)
            month = int(row.get("month", 0) or 0)

            all_records.append({
                "연도":      year,
                "월":        month,
                "연월":      f"{year}{month:02d}",
                "지역":      branch.get("entity_name", ""),
                "영업점":    branch.get("branch_name", ""),
                "매출":          revenue,
                "재료비":        mat,
                "재료비율":      mat / m,
                "인건비":        labor,
                "인건비율":      labor / m,
                "경비":          exp,
                "경비율":        exp / m,
                "매출총이익":    gross,
                "매출총이익율":  gross / m,
                "법인비용":      corp,
                "영업이익":      op,
                "영업이익율":    op / m,
                "운영팀인건비":    ops_labor,
                "운영팀인건비율":  ops_labor / m,
                "운영팀경비":      ops_exp,
                "운영팀경비율":    ops_exp / m,
                "사업부공통비":    div_cost,
                "사업부공통비율":  div_cost / m,
                "영업이익Ⅱ":      op2,
                "영업이익Ⅱ율":    op2 / m,
            })

        if len(rows) < page_size:
            break
        offset += page_size

    return all_records


# =====================
# 서울외국환중개 XML
# =====================
def fetch_smbs_xml(endpoint, currency, start, end, referer=None):
    arr_value = f"{currency}_{start}_{end}"
    headers = {**SMBS_HEADERS}
    if referer:
        headers["Referer"] = referer
    try:
        res = requests.get(
            f"http://www.smbs.biz/ExRate/{endpoint}?arr_value={arr_value}",
            headers=headers, timeout=10, verify=False
        )
        content = res.content.decode("euc-kr").strip()
        pattern = re.compile(r"<set[^>]+label='([^']+)'[^>]+value='([^']+)'")
        data = {}
        for match in pattern.finditer(content):
            label = match.group(1).strip()
            value = match.group(2).strip()
            if len(label.split(".")[0]) == 2:
                parts = label.split(".")
                key = f"20{parts[0]}{parts[1]}{parts[2]}"
            else:
                key = label.replace(".", "")
            data[key] = value
        return data
    except Exception as e:
        print(f"SMBS XML 오류 ({currency}): {e}")
        return {}

def to_dash(date_str):
    return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"

def get_latest_date():
    date = datetime.now()
    for _ in range(10):
        date_str  = date.strftime("%Y%m%d")
        formatted = to_dash(date_str)
        data = fetch_smbs_xml(
            "StdExRate_xml.jsp", "USD", formatted, formatted,
            "http://www.smbs.biz/ExRate/StdExRate.jsp"
        )
        if data:
            return date_str
        date -= timedelta(days=1)
    return datetime.now().strftime("%Y%m%d")

def fetch_smbs_today(currency, date_str):
    formatted = to_dash(date_str)
    data = fetch_smbs_xml(
        "StdExRate_xml.jsp", currency, formatted, formatted,
        "http://www.smbs.biz/ExRate/StdExRate.jsp"
    )
    if date_str in data:
        return data[date_str]
    if data:
        return list(data.values())[-1]
    return ""

def fetch_smbs_monthly_avg(currency, year, month):
    data = fetch_smbs_xml(
        "MonAvgStdExRate_xml.jsp", currency,
        f"{year-1}-{month:02d}", f"{year}-{month:02d}",
        "http://www.smbs.biz/ExRate/MonAvgStdExRate.jsp"
    )
    target_key = f"{year}{month:02d}"
    if target_key not in data and data:
        target_key = sorted(data.keys())[-1]
    return data.get(target_key, "")

def fetch_smbs_month_end(currency, year, month):
    data = fetch_smbs_xml(
        "MonLastStdExRate_xml.jsp", currency,
        f"{year-1}-{month:02d}", f"{year}-{month:02d}",
        "http://www.smbs.biz/ExRate/MonLastStdExRate.jsp"
    )
    target_key = f"{year}{month:02d}"
    if target_key not in data and data:
        target_key = sorted(data.keys())[-1]
    return data.get(target_key, "")

def calc_change(today_val, yesterday_val, decimal=2):
    try:
        t = float(str(today_val).replace(",", ""))
        y = float(str(yesterday_val).replace(",", ""))
        diff = t - y
        if diff > 0:
            return "RISE", f"{diff:+.{decimal}f}"
        elif diff < 0:
            return "FALL", f"{diff:+.{decimal}f}"
        else:
            return "EVEN", "0.00"
    except:
        return "", ""

def fetch_er_open():
    try:
        res  = requests.get("https://open.er-api.com/v6/latest/KRW", timeout=10, verify=False)
        data = res.json().get("rates", {})
        result = {}
        for cur in ["IQD", "LBP"]:
            if cur in data and data[cur] != 0:
                result[cur] = str(round(1 / data[cur], 4))
        return result
    except Exception as e:
        print(f"ER Open 오류: {e}")
        return {}


# =====================
# 기본
# =====================
@app.get("/")
def root():
    return {"status": "ok", "message": "환율 API 서버 작동중"}

# =====================
# 환율
# =====================
@app.get("/rates")
def get_rates():
    try:
        cached = get_cached_rates()
        if cached:
            return cached

        today_str = get_latest_date()
        prev_date = datetime.strptime(today_str, "%Y%m%d") - timedelta(days=1)
        yesterday_str = ""
        for _ in range(10):
            ps = prev_date.strftime("%Y%m%d")
            if fetch_smbs_today("USD", ps):
                yesterday_str = ps
                break
            prev_date -= timedelta(days=1)

        rates = []
        for cur in ALL_TARGET:
            try:
                today_val     = fetch_smbs_today(cur, today_str)
                yesterday_val = fetch_smbs_today(cur, yesterday_str) if yesterday_str else ""
                if today_val:
                    decimal = 4 if cur == "KZT" else 2
                    change, change_val = calc_change(today_val, yesterday_val, decimal)
                    rates.append({
                        "currency": cur, "name": CUR_NAMES[cur],
                        "base": today_val, "buy": "-", "sell": "-",
                        "change": change, "change_val": change_val,
                    })
            except Exception as e:
                print(f"{cur} 오류: {e}")

        try:
            er_data = fetch_er_open()
            for cur in ["IQD", "LBP"]:
                if cur in er_data:
                    rates.append({
                        "currency": cur, "name": CUR_NAMES[cur],
                        "base": er_data[cur], "buy": "-", "sell": "-",
                        "change": "", "change_val": "",
                    })
        except Exception as e:
            print(f"IQD/LBP 오류: {e}")

        result = {
            "success": True,
            "date": today_str,
            "updated_at": datetime.now().strftime("%H:%M:%S"),
            "data": rates
        }
        set_rates_cache(result)
        return result

    except Exception as e:
        return {"success": False, "error": str(e), "data": []}

@app.get("/rates/by-date")
def get_rates_by_date(date: str):
    try:
        prev_date = datetime.strptime(date, "%Y%m%d") - timedelta(days=1)
        prev_str  = ""
        for _ in range(10):
            ps = prev_date.strftime("%Y%m%d")
            if fetch_smbs_today("USD", ps):
                prev_str = ps
                break
            prev_date -= timedelta(days=1)

        rates = []
        for cur in ALL_TARGET:
            try:
                today_val     = fetch_smbs_today(cur, date)
                yesterday_val = fetch_smbs_today(cur, prev_str) if prev_str else ""
                if today_val:
                    decimal = 4 if cur == "KZT" else 2
                    change, change_val = calc_change(today_val, yesterday_val, decimal)
                    rates.append({
                        "currency": cur, "name": CUR_NAMES[cur],
                        "base": today_val, "buy": "-", "sell": "-",
                        "change": change, "change_val": change_val,
                    })
            except Exception as e:
                print(f"{cur} 오류: {e}")

        try:
            er_data = fetch_er_open()
            for cur in ["IQD", "LBP"]:
                if cur in er_data:
                    rates.append({
                        "currency": cur, "name": CUR_NAMES[cur],
                        "base": er_data[cur], "buy": "-", "sell": "-",
                        "change": "", "change_val": "",
                    })
        except:
            pass

        if not rates:
            return {"success": False, "error": "데이터 없음 (주말/공휴일)", "data": []}

        return {
            "success": True, "date": date,
            "updated_at": datetime.now().strftime("%H:%M:%S"),
            "data": rates
        }
    except Exception as e:
        return {"success": False, "error": str(e), "data": []}

@app.get("/rates/monthly-avg")
def get_monthly_avg(year: int, month: int):
    try:
        result = []
        for cur in ALL_TARGET:
            try:
                val = fetch_smbs_monthly_avg(cur, year, month)
                if val:
                    result.append({"currency": cur, "name": CUR_NAMES[cur],
                                   "base": val, "buy": "-", "sell": "-"})
            except Exception as e:
                print(f"{cur} 월평균 오류: {e}")
        return {"success": True, "year": year, "month": month, "data": result}
    except Exception as e:
        return {"success": False, "error": str(e), "data": []}

@app.get("/rates/month-end")
def get_month_end(year: int, month: int):
    try:
        result = []
        for cur in ALL_TARGET:
            try:
                val = fetch_smbs_month_end(cur, year, month)
                if val:
                    result.append({"currency": cur, "name": CUR_NAMES[cur],
                                   "base": val, "buy": "-", "sell": "-"})
            except Exception as e:
                print(f"{cur} 월말 오류: {e}")
        return {"success": True, "year": year, "month": month, "data": result}
    except Exception as e:
        return {"success": False, "error": str(e), "data": []}

@app.get("/rates/weekly")
def get_weekly(currency: str = "USD"):
    try:
        result = []
        date   = datetime.now()
        count  = 0
        while count < 15:
            date_str = date.strftime("%Y%m%d")
            val = fetch_smbs_today(currency, date_str)
            if val:
                result.append({
                    "date": f"{date.month}/{date.day}",
                    "value": float(val.replace(",", "")),
                    "full_date": date_str,
                })
                count += 1
            date -= timedelta(days=1)
        result.reverse()
        return {"success": True, "currency": currency, "data": result}
    except Exception as e:
        return {"success": False, "error": str(e), "data": []}

@app.get("/rates/cache/status")
def rates_cache_status():
    if "timestamp" not in _rates_cache:
        return {"cached": False}
    age = int(time.time() - _rates_cache["timestamp"])
    return {"cached": True, "age_seconds": age, "expires_in": max(0, RATES_CACHE_TTL - age)}

@app.post("/rates/cache/clear")
def clear_rates_cache():
    _rates_cache.clear()
    return {"success": True, "message": "환율 캐시 초기화 완료"}

@app.get("/debug/today")
def debug_today():
    try:
        today_str = get_latest_date()
        result    = {cur: fetch_smbs_today(cur, today_str) for cur in ALL_TARGET}
        return {"today": today_str, "formatted": to_dash(today_str), "result": result}
    except Exception as e:
        return {"error": str(e)}

# =====================
# 인증
# =====================
@app.post("/auth/register")
def register(data: dict):
    try:
        email       = data.get("email", "").strip()
        name        = data.get("name", "").strip()
        employee_id = data.get("employee_id", "").strip()
        password    = data.get("password", "").strip()

        if not all([email, name, employee_id, password]):
            return {"success": False, "error": "모든 항목을 입력해주세요."}
        if supabase.table("users").select("id").eq("employee_id", employee_id).execute().data:
            return {"success": False, "error": "이미 등록된 사번입니다."}
        if supabase.table("users").select("id").eq("email", email).execute().data:
            return {"success": False, "error": "이미 등록된 이메일입니다."}

        hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        supabase.table("users").insert({
            "email": email, "name": name,
            "employee_id": employee_id, "password": hashed, "approved": True,
        }).execute()
        return {"success": True, "message": "계정 신청이 완료됐습니다. 관리자 승인 후 로그인 가능합니다."}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/auth/login")
def login(data: dict):
    try:
        employee_id = data.get("employee_id", "").strip()
        password    = data.get("password", "").strip()

        if not employee_id or not password:
            return {"success": False, "error": "사번과 비밀번호를 입력해주세요."}

        result = supabase.table("users").select("*").eq("employee_id", employee_id).execute()
        if not result.data:
            return {"success": False, "error": "존재하지 않는 사번입니다."}

        user = result.data[0]
        if not user.get("approved"):
            return {"success": False, "error": "관리자 승인 대기 중입니다."}
        if not bcrypt.checkpw(password.encode("utf-8"), user["password"].encode("utf-8")):
            return {"success": False, "error": "비밀번호가 올바르지 않습니다."}

        return {
            "success": True,
            "user": {
                "employee_id": user["employee_id"],
                "name": user["name"],
                "email": user["email"],
                "role": user.get("role", "user"),
            }
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/auth/pending")
def get_pending():
    try:
        result = supabase.table("users").select("*").eq("approved", False).execute()
        return {"success": True, "data": result.data}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/auth/approve")
def approve_user(data: dict):
    try:
        employee_id = data.get("employee_id")
        supabase.table("users").update({"approved": True}).eq("employee_id", employee_id).execute()
        return {"success": True, "message": f"{employee_id} 승인 완료"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/auth/reject")
def reject_user(data: dict):
    try:
        employee_id = data.get("employee_id")
        supabase.table("users").delete().eq("employee_id", employee_id).execute()
        return {"success": True, "message": f"{employee_id} 거절 완료"}
    except Exception as e:
        return {"success": False, "error": str(e)}

# =====================
# PnL 캐시 관리
# =====================
@app.get("/pnl/cache/status")
def pnl_cache_status():
    if "timestamp" not in _pnl_cache:
        return {"cached": False}
    age = int(time.time() - _pnl_cache["timestamp"])
    return {
        "cached": True,
        "records": len(_pnl_cache.get("records", [])),
        "age_seconds": age,
        "expires_in": max(0, PNL_CACHE_TTL - age)
    }

@app.post("/pnl/cache/clear")
def clear_pnl_cache():
    _pnl_cache.clear()
    return {"success": True, "message": "PnL 캐시 초기화 완료"}

# =====================
# PnL 엔드포인트 (Framer가 호출하는 API — 필드명 동일 유지)
# =====================
@app.get("/pnl/raw")
def get_raw_data(year: int = None, month: int = None, region: str = None, store: str = None):
    try:
        records = get_cached_records()
        if year:    records = [r for r in records if r["연도"] == year]
        if month:   records = [r for r in records if r["월"] == month]
        if region:  records = [r for r in records if r["지역"] == region]
        if store:   records = [r for r in records if store in r["영업점"]]
        return {"success": True, "count": len(records), "data": records}
    except Exception as e:
        return {"success": False, "error": str(e), "data": []}

@app.get("/pnl/monthly")
def get_monthly(year: int, month: int):
    try:
        records = get_cached_records()
        records = [r for r in records if r["연도"] == year and r["월"] == month]

        region_summary = {}
        for r in records:
            region = r["지역"]
            if region not in region_summary:
                region_summary[region] = {
                    "지역": region, "매출": 0, "재료비": 0, "인건비": 0,
                    "경비": 0, "매출총이익": 0, "법인비용": 0,
                    "영업이익": 0, "영업점_수": 0, "영업점목록": []
                }
            s = region_summary[region]
            for key in ["매출","재료비","인건비","경비","매출총이익","법인비용","영업이익"]:
                s[key] += r[key]
            s["영업점_수"] += 1
            s["영업점목록"].append(r)

        for region, s in region_summary.items():
            m = s["매출"]
            if m > 0:
                s["재료비율"]    = f"{s['재료비']/m*100:.2f}%"
                s["인건비율"]    = f"{s['인건비']/m*100:.2f}%"
                s["경비율"]      = f"{s['경비']/m*100:.2f}%"
                s["매출총이익율"] = f"{s['매출총이익']/m*100:.2f}%"
                s["영업이익율"]  = f"{s['영업이익']/m*100:.2f}%"
            else:
                s["재료비율"] = s["인건비율"] = s["경비율"] = s["매출총이익율"] = s["영업이익율"] = "0.00%"

        return {"success": True, "year": year, "month": month, "data": list(region_summary.values())}
    except Exception as e:
        return {"success": False, "error": str(e), "data": []}

@app.get("/pnl/cumulative")
def get_cumulative(year: int, start_month: int = 1, end_month: int = 12):
    try:
        records = get_cached_records()
        records = [r for r in records if r["연도"]==year and start_month<=r["월"]<=end_month]

        total  = {"매출":0,"재료비":0,"인건비":0,"경비":0,"매출총이익":0,"법인비용":0,"영업이익":0}
        region_summary = {}

        for r in records:
            for key in total: total[key] += r[key]
            region = r["지역"]
            if region not in region_summary:
                region_summary[region] = {
                    "지역": region, "매출":0,"재료비":0,"인건비":0,
                    "경비":0,"매출총이익":0,"법인비용":0,"영업이익":0
                }
            for key in ["매출","재료비","인건비","경비","매출총이익","법인비용","영업이익"]:
                region_summary[region][key] += r[key]

        def add_rates(s):
            m = s["매출"]
            if m > 0:
                s["재료비율"]    = f"{s['재료비']/m*100:.2f}%"
                s["인건비율"]    = f"{s['인건비']/m*100:.2f}%"
                s["경비율"]      = f"{s['경비']/m*100:.2f}%"
                s["매출총이익율"] = f"{s['매출총이익']/m*100:.2f}%"
                s["영업이익율"]  = f"{s['영업이익']/m*100:.2f}%"
            else:
                s["재료비율"] = s["인건비율"] = s["경비율"] = s["매출총이익율"] = s["영업이익율"] = "0.00%"
            return s

        add_rates(total)
        for region in region_summary: add_rates(region_summary[region])

        return {
            "success": True, "year": year,
            "start_month": start_month, "end_month": end_month,
            "total": total, "by_region": list(region_summary.values())
        }
    except Exception as e:
        return {"success": False, "error": str(e), "data": []}

@app.get("/pnl/regions")
def get_regions():
    try:
        records = get_cached_records()
        regions = sorted(list(set(r["지역"] for r in records if r["지역"])))
        return {"success": True, "data": regions}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/pnl/stores")
def get_stores(region: str = None):
    try:
        records = get_cached_records()
        if region:
            records = [r for r in records if r["지역"] == region]
        stores = sorted(list(set(r["영업점"] for r in records if r["영업점"])))
        return {"success": True, "data": stores}
    except Exception as e:
        return {"success": False, "error": str(e)}

# =====================
# 엑셀 양식 다운로드
# =====================
@app.get("/pnl/template")
def download_template():
    """업로드용 빈 양식 엑셀 다운로드"""
    try:
        # branches 목록 조회
        branches_res = supabase.table("branches").select(
            "id, branch_name, entity_name"
        ).eq("is_active", True).execute()
        branches = branches_res.data or []

        wb = Workbook()
        ws = wb.active
        ws.title = "손익입력"

        # 헤더
        headers = ["지역", "영업점", "연도", "월",
                   "매출", "재료비", "인건비", "경비",
                   "법인비용", "매출총이익", "영업이익"]
        ws.append(headers)

        # 헤더 스타일
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1E293B")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 영업점 목록 미리 채워넣기
        for b in branches:
            ws.append([
                b["entity_name"],  # 지역
                b["branch_name"],  # 영업점
                "",  # 연도 (입력)
                "",  # 월 (입력)
                "", "", "", "", "", "", ""  # 수치 (입력)
            ])

        # 열 너비
        col_widths = [10, 20, 8, 6, 12, 12, 12, 12, 12, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=pnl_template.xlsx"}
        )
    except Exception as e:
        return {"success": False, "error": str(e)}


# =====================
# Raw Data 엑셀 다운로드
# =====================
@app.get("/pnl/export")
def export_pnl(year: int = None, month: int = None, region: str = None):
    """현재 필터 기준 손익 데이터 엑셀 다운로드"""
    try:
        records = get_cached_records()
        if year:   records = [r for r in records if r["연도"] == year]
        if month:  records = [r for r in records if r["월"] == month]
        if region: records = [r for r in records if r["지역"] == region]

        # 정렬: 지역 → 영업점 → 연도 → 월
        records = sorted(records, key=lambda r: (r["지역"], r["영업점"], r["연도"], r["월"]))

        wb = Workbook()
        ws = wb.active
        ws.title = "손익데이터"

        headers = ["지역", "영업점", "연도", "월",
                   "매출", "재료비", "재료비율",
                   "인건비", "인건비율",
                   "경비", "경비율",
                   "매출총이익", "매출총이익율",
                   "법인비용", "영업이익", "영업이익율"]
        ws.append(headers)

        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1E293B")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        def pct(v):
            return f"{v*100:.1f}%"

        for r in records:
            ws.append([
                r["지역"], r["영업점"], r["연도"], r["월"],
                r["매출"], r["재료비"], pct(r["재료비율"]),
                r["인건비"], pct(r["인건비율"]),
                r["경비"], pct(r["경비율"]),
                r["매출총이익"], pct(r["매출총이익율"]),
                r["법인비용"], r["영업이익"], pct(r["영업이익율"]),
            ])

        # 열 너비
        col_widths = [10, 20, 8, 6, 12, 12, 10, 12, 10, 12, 10, 14, 12, 12, 14, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"pnl_export_{year or 'all'}_{month or 'all'}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return {"success": False, "error": str(e)}


# =====================
# 엑셀 업로드 → Supabase 저장
# =====================
@app.post("/pnl/upload")
async def upload_pnl(file: UploadFile = File(...)):
    """
    엑셀 업로드 → pnl_monthly 테이블에 upsert
    컬럼: 지역, 영업점, 연도, 월, 매출, 재료비, 인건비, 경비, 법인비용, 매출총이익, 영업이익
    """
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active

        # branches 맵 (branch_name → id)
        branches_res = supabase.table("branches").select("id, branch_name, entity_name").execute()
        branch_map = {b["branch_name"]: b["id"] for b in (branches_res.data or [])}

        rows_to_upsert = []
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # 빈 행 스킵
                continue

            지역, 영업점, 연도, 월, 매출, 재료비, 인건비, 경비, 법인비용, 매출총이익, 영업이익 = (
                row[0], row[1], row[2], row[3], row[4],
                row[5], row[6], row[7], row[8], row[9], row[10]
            )

            # 영업점 → branch_id 매핑
            branch_id = branch_map.get(str(영업점).strip())
            if not branch_id:
                errors.append(f"행 {row_idx}: '{영업점}' 영업점을 찾을 수 없음")
                continue

            if not 연도 or not 월:
                errors.append(f"행 {row_idx}: 연도/월 누락")
                continue

            rows_to_upsert.append({
                "branch_id":        branch_id,
                "year":             int(연도),
                "month":            int(월),
                "revenue":          float(매출 or 0),
                "material_cost":    float(재료비 or 0),
                "labor_cost":       float(인건비 or 0),
                "expenses":         float(경비 or 0),
                "hq_allocated_cost": float(법인비용 or 0),
                "gross_profit":     float(매출총이익 or 0),
                "operating_profit": float(영업이익 or 0),
            })

        if not rows_to_upsert:
            return {"success": False, "error": "저장할 데이터가 없습니다.", "errors": errors}

        # upsert (branch_id + year + month 기준 중복 처리)
        supabase.table("pnl_monthly").upsert(
            rows_to_upsert,
            on_conflict="branch_id,year,month"
        ).execute()

        # 캐시 초기화 (새 데이터 반영)
        _pnl_cache.clear()

        return {
            "success": True,
            "inserted": len(rows_to_upsert),
            "errors": errors,
            "message": f"{len(rows_to_upsert)}건 저장 완료"
        }

    except Exception as e:
        return {"success": False, "error": str(e)}

from fastapi import UploadFile, File, Form
from io import BytesIO
import openpyxl
import pandas as pd

# ── 행번호 → 영업점명 매핑 ──────────────────────────
ROW_BRANCH_MAP = {
    10: "일조 위아",      11: "일조 파워텍",
    12: "SKON",           13: "장가항 위아",
    14: "무석 하이닉스",  15: "광저우 HTWO",
    16: "랑방 오리온",    18: "북경 기차빌딩",
    19: "북경 만도연구소",20: "천진 모비스",
    21: "연태 현대차",    23: "상해 케미컬",
    24: "상해 모비스",    25: "상해 엘리베이터",
    26: "중경 하이닉스",  27: "상숙 컨티넨탈",
    28: "양중 중공업",    29: "무석 모비스",
    30: "염성 모비스",    32: "몬테레이",
    33: "트랜스리드",     34: "DBNR",
    35: "핫도그 공장",    37: "조지아",
    38: "서배너",         39: "HSAGP(SKON)",
    41: "마잔",           42: "자프라",
    44: "바스라",         46: "현대건설BNPP",
    47: "아크부대",       49: "동명부대",
}

# ── Raw 파일 컬럼 매핑 ──────────────────────────────
RAW1_ACTL_COLS = {
    "revenue":          17,  # G열
    "material_cost":    30,  # L열
    "labor_cost":       36,  # R열
    "expenses":         42,  # X열
    "hq_allocated_cost":58,  # AG열
}

RAW1_ROW_MAP = {
    10:[16], 11:[17], 12:[21], 13:[18], 14:[19], 15:[20], 16:[22],
    18:[25], 19:[28], 20:[26], 21:[27],
    23:[31], 24:[34], 25:[35], 26:[32], 27:[33], 28:[36], 29:[41],
    30:[37,38,39,40],
    32:[44,45,46,47], 33:[48], 34:[50], 35:[49],
    37:[52], 38:[53], 39:[54],
}
RAW1_PREV_ONLY_MAP = {
    41:[11], 42:[12], 44:[13], 46:[9], 47:[10], 49:[14],
}

# ── GL 파일 설정 ──────────────────────────────────
GL_VALUE_COL   = 10
GL_DEPT_COL    = 7
GL_SUBJECT_COL = 2
GL_ITEM_COL    = 3
UNIT_DIV       = 1_000_000

GL_SALES_SUBJ    = ('제품매출', '용역매출')
GL_MATERIAL_SUBJ = ('원재료',)
GL_LABOR_SUBJ    = ('노무비',)
GL_EXPENSE_SUBJ  = ('경비',)
GL_JISA_SUBJ     = ('일반관리비',)
GL_EXCLUDE_ITEM  = ('대체',)
GL_EXCLUDE_SEKMOK = ('년차수당(발생분)', '년차수당(미소진분)', '근속포상금')
GL_RECLASSIFY_TO_LABOR = ('용역비',)

GL_DEPT_ROW_MAP = {
    'Marjan': 41, 'Jafurah': 42, 'Basrah': 44,
    '현대건설BNPP': 46, '특전사아크부대': 47, '레바논 동명부대': 49,
}
JISA_PROP_MAP = {
    '사우디지사': ['Marjan', 'Jafurah'],
    'UAE지사':   ['현대건설BNPP', '특전사아크부대'],
}
JISA_FIXED_MAP = {
    '이라크지사': [44],
    '레바논지사': [49],
}

# ── 법인 파일 설정 ────────────────────────────────
CORP_US_SHEETS = {
    '조지아-월별 손익 _ 원화': 37,
    '서배너-월별 손익 _ 원화': 38,
    'HSAGP-월별 손익 _ 원화': 39,
}
CORP_CN1_SHEETS = {
    '위아':10, '파워텍':11, '염성계':12, '장가항계':13,
    '무석하이닉스계':14, '광저우계':15, '랑방계':16,
}
CORP_CN2_SHEETS = {
    '식당':18, '북경만도연구소':19, '천진계':20, '연태계':21,
}
CORP_CN3_SHEETS = {
    '韩华化学':23, '上海摩比斯店':24, '상해현대엘리베이터':25,
    '중경계':26, '상숙계':27, '양중계':28, '무석계':29, '염성계':30,
}

CORP_DATA_ROW_COL = {8:'revenue', 15:'material_cost', 22:'labor_cost', 29:'expenses'}
CORP_JISA_ROW     = 52
CORP_JISA_COL     = 6

CORP_MX_SHEET  = '3. 당월손익_기장기준 + 조정내역(원화환산)'
CORP_MX_MAP    = {32:5, 33:11, 35:14}
CORP_MX_ROWS   = {9:'revenue', 16:'material_cost', 23:'labor_cost', 30:'expenses'}
CORP_MX_JISA   = (53, 18)


# ── 공통: branch_name → branch_id 매핑 ──────────
def get_branch_map():
    res = supabase.table("branches").select("id, branch_name").execute()
    return {b["branch_name"]: b["id"] for b in (res.data or [])}


def row_data_to_supabase(row_lookup: dict, yr: int, mo: int) -> list:
    """row_num → {field: value} 를 Supabase upsert 형식으로 변환"""
    branch_map = get_branch_map()
    rows = []
    for row_num, fields in row_lookup.items():
        branch_name = ROW_BRANCH_MAP.get(row_num)
        if not branch_name:
            continue
        branch_id = branch_map.get(branch_name)
        if not branch_id:
            print(f"[WARN] branch_id 없음: {branch_name}")
            continue
        rev  = float(fields.get("revenue", 0) or 0)
        mat  = float(fields.get("material_cost", 0) or 0)
        lab  = float(fields.get("labor_cost", 0) or 0)
        exp  = float(fields.get("expenses", 0) or 0)
        corp = float(fields.get("hq_allocated_cost", 0) or 0)
        rows.append({
            "branch_id":         branch_id,
            "year":              yr,
            "month":             mo,
            "revenue":           round(rev, 3),
            "material_cost":     round(mat, 3),
            "labor_cost":        round(lab, 3),
            "expenses":          round(exp, 3),
            "hq_allocated_cost": round(corp, 3),
            "gross_profit":      round(rev - mat - lab - exp, 3),
            "operating_profit":  round(rev - mat - lab - exp - corp, 3),
        })
    return rows


# ── 파서: Raw 영업점별 요약 ──────────────────────
def parse_raw_file(content: bytes) -> dict:
    df = pd.read_excel(BytesIO(content), sheet_name='당월', header=None)
    result = {}

    def _sum(col_idx, excel_rows):
        total = 0.0
        for er in excel_rows:
            try:
                val = df.iloc[er - 1, col_idx]
                total += float(val) if pd.notna(val) else 0.0
            except:
                pass
        return total / 1000

    for row_num, excel_rows in {**RAW1_ROW_MAP, **RAW1_PREV_ONLY_MAP}.items():
        result[row_num] = {
            field: _sum(col_idx, excel_rows)
            for field, col_idx in RAW1_ACTL_COLS.items()
        }
    return result


# ── 파서: GL 중동 ────────────────────────────────
def parse_gl_file(content: bytes) -> dict:
    df = pd.read_excel(BytesIO(content), header=None)
    mask = (
        ~df[GL_SUBJECT_COL].astype(str).str.contains('합계', na=True) &
        ~df[GL_ITEM_COL].astype(str).str.contains('합계', na=True)
    )
    data = df[mask].iloc[2:].copy()
    data = data[~data[GL_ITEM_COL].isin(GL_EXCLUDE_ITEM)]
    data = data[~data[4].isin(GL_EXCLUDE_SEKMOK)]
    result = {}

    def _add(row_num, field, val):
        if row_num not in result: result[row_num] = {}
        result[row_num][field] = result[row_num].get(field, 0.0) + val / UNIT_DIV

    for dept, row_num in GL_DEPT_ROW_MAP.items():
        sub = data[data[GL_DEPT_COL] == dept]
        _add(row_num, "revenue",       sub[sub[GL_SUBJECT_COL].isin(GL_SALES_SUBJ)][GL_VALUE_COL].sum())
        _add(row_num, "material_cost", sub[sub[GL_SUBJECT_COL].isin(GL_MATERIAL_SUBJ)][11].sum())
        labor = sub[sub[GL_SUBJECT_COL].isin(GL_LABOR_SUBJ)]
        reclassify = sub[(sub[GL_SUBJECT_COL].isin(GL_EXPENSE_SUBJ)) &
                         (sub[GL_ITEM_COL].isin(GL_RECLASSIFY_TO_LABOR))]
        _add(row_num, "labor_cost", labor[GL_VALUE_COL].sum() + reclassify[GL_VALUE_COL].sum())
        expense = sub[(sub[GL_SUBJECT_COL].isin(GL_EXPENSE_SUBJ)) &
                      (~sub[GL_ITEM_COL].isin(GL_RECLASSIFY_TO_LABOR))]
        _add(row_num, "expenses", expense[GL_VALUE_COL].sum())

    for jisa_dept, shop_depts in JISA_PROP_MAP.items():
        sub = data[(data[GL_DEPT_COL] == jisa_dept) &
                   (data[GL_SUBJECT_COL].isin(GL_JISA_SUBJ))]
        total = sub[GL_VALUE_COL].sum() / UNIT_DIV
        if total == 0: continue
        sales = {}
        for dept in shop_depts:
            rn = GL_DEPT_ROW_MAP.get(dept)
            if rn is None: continue
            s = data[data[GL_DEPT_COL] == dept]
            sales[rn] = s[s[GL_SUBJECT_COL].isin(GL_SALES_SUBJ)][GL_VALUE_COL].sum()
        total_sales = sum(sales.values())
        for rn, sale in sales.items():
            ratio = (sale / total_sales) if total_sales != 0 else (1 / len(sales))
            if rn not in result: result[rn] = {}
            result[rn]["hq_allocated_cost"] = result[rn].get("hq_allocated_cost", 0.0) + total * ratio

    for jisa_dept, target_rows in JISA_FIXED_MAP.items():
        sub = data[(data[GL_DEPT_COL] == jisa_dept) &
                   (data[GL_SUBJECT_COL].isin(GL_JISA_SUBJ))]
        total = sub[GL_VALUE_COL].sum() / UNIT_DIV
        for rn in target_rows:
            if rn not in result: result[rn] = {}
            result[rn]["hq_allocated_cost"] = result[rn].get("hq_allocated_cost", 0.0) + total

    return result


# ── 파서: 법인 공통 (산동/북경/상해) ─────────────
def parse_corp_cn(content: bytes, sheet_row_map: dict, jisa_sheet: str) -> dict:
    result = {}
    for sheet_name, row_num in sheet_row_map.items():
        try:
            df = pd.read_excel(BytesIO(content), sheet_name=sheet_name, header=None)
            result[row_num] = {}
            for data_row, field in CORP_DATA_ROW_COL.items():
                try:
                    val = float(df.iloc[data_row - 1, CORP_JISA_COL - 1] or 0) / 1000
                except:
                    val = 0.0
                result[row_num][field] = val
        except:
            continue

    try:
        df_j = pd.read_excel(BytesIO(content), sheet_name=jisa_sheet, header=None)
        jisa_total = abs(float(df_j.iloc[CORP_JISA_ROW - 1, CORP_JISA_COL - 1] or 0) / 1000)
        if jisa_total != 0:
            total_sales = sum(result[r].get("revenue", 0) for r in result)
            for r in result:
                ratio = (result[r].get("revenue", 0) / total_sales) if total_sales != 0 else (1 / len(result))
                result[r]["hq_allocated_cost"] = jisa_total * ratio
    except:
        pass
    return result


# ── 파서: 미국법인 ───────────────────────────────
def parse_corp_us(content: bytes, mo: int) -> dict:
    mo_col_idx = 3 + mo
    DATA_ROW_COL = {8:"revenue", 15:"material_cost", 22:"labor_cost", 29:"expenses", 54:"hq_allocated_cost"}
    result = {}
    for sheet_name, row_num in CORP_US_SHEETS.items():
        try:
            df = pd.read_excel(BytesIO(content), sheet_name=sheet_name, header=None)
            result[row_num] = {}
            for data_row, field in DATA_ROW_COL.items():
                try:
                    val = float(df.iloc[data_row - 1, mo_col_idx] or 0) / 1000
                except:
                    val = 0.0
                result[row_num][field] = val
        except:
            continue
    return result


# ── 파서: 멕시코법인 ─────────────────────────────
def parse_corp_mx(content: bytes) -> dict:
    result = {}
    try:
        df = pd.read_excel(BytesIO(content), sheet_name=CORP_MX_SHEET, header=None)
        for row_num, excel_col in CORP_MX_MAP.items():
            result[row_num] = {}
            for data_row, field in CORP_MX_ROWS.items():
                try:
                    val = float(df.iloc[data_row - 1, excel_col - 1] or 0) / 1000
                except:
                    val = 0.0
                result[row_num][field] = val
        try:
            jisa_total = abs(float(df.iloc[CORP_MX_JISA[0]-1, CORP_MX_JISA[1]-1] or 0) / 1000)
            if jisa_total != 0:
                total_sales = sum(result[r].get("revenue", 0) for r in result)
                for r in result:
                    ratio = (result[r].get("revenue", 0) / total_sales) if total_sales != 0 else (1/len(result))
                    result[r]["hq_allocated_cost"] = jisa_total * ratio
        except:
            pass
    except:
        pass
    return result


# ── 통합 업로드 엔드포인트 ───────────────────────
@app.post("/pnl/upload-source")
async def upload_source(
    year:     int        = Form(...),
    month:    int        = Form(...),
    raw:      UploadFile = File(None),
    gl:       UploadFile = File(None),
    corp_us:  UploadFile = File(None),
    corp_cn1: UploadFile = File(None),
    corp_cn2: UploadFile = File(None),
    corp_cn3: UploadFile = File(None),
    corp_mx:  UploadFile = File(None),
):
    try:
        row_lookup = {}

        def _merge(parsed: dict):
            for rn, fields in parsed.items():
                if rn not in row_lookup: row_lookup[rn] = {}
                for f, v in fields.items():
                    row_lookup[rn][f] = row_lookup[rn].get(f, 0.0) + v

        if raw:
            _merge(parse_raw_file(await raw.read()))
        if gl:
            _merge(parse_gl_file(await gl.read()))
        if corp_us:
            _merge(parse_corp_us(await corp_us.read(), month))
        if corp_cn1:
            content = await corp_cn1.read()
            _merge(parse_corp_cn(content, CORP_CN1_SHEETS, '복덕찬음'))
        if corp_cn2:
            content = await corp_cn2.read()
            _merge(parse_corp_cn(content, CORP_CN2_SHEETS, '본사'))
        if corp_cn3:
            content = await corp_cn3.read()
            _merge(parse_corp_cn(content, CORP_CN3_SHEETS, '본사'))
        if corp_mx:
            _merge(parse_corp_mx(await corp_mx.read()))

        if not row_lookup:
            return {"success": False, "error": "업로드된 파일이 없습니다."}

        rows = row_data_to_supabase(row_lookup, year, month)
        if not rows:
            return {"success": False, "error": "변환된 데이터가 없습니다."}

        supabase.table("pnl_monthly").upsert(
            rows, on_conflict="branch_id,year,month"
        ).execute()

        _pnl_cache.clear()

        return {
            "success":  True,
            "year":     year,
            "month":    month,
            "inserted": len(rows),
            "message":  f"{year}년 {month}월 {len(rows)}개 영업점 저장 완료"
        }

    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/pnl/debug-sheets")
async def debug_sheets(file: UploadFile = File(...)):
    """업로드된 엑셀 파일의 시트 목록 반환"""
    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
        return {"success": True, "sheets": wb.sheetnames}
    except Exception as e:
        return {"success": False, "error": str(e)}

from pydantic import BaseModel
from fastapi import Header, HTTPException

# 1. 요청 데이터를 받기 위한 모델 정의
class UpdatePasswordRequest(BaseModel):
    currentPassword: str
    newPassword: str

# 2. 비밀번호 변경 엔드포인트
@app.post("/user/update")
async def update_password(
    request: UpdatePasswordRequest, 
    authorization: str = Header(None)
):
    if not authorization:
        raise HTTPException(status_code=401, detail="인증 토큰이 없습니다.")
    
    # Bearer 토큰 추출
    token = authorization.replace("Bearer ", "")
    
    try:
        # ① 현재 토큰으로 유저 정보 가져오기
        user = supabase.auth.get_user(token)
        if not user:
            raise HTTPException(status_code=401, detail="유효하지 않은 세션입니다.")
        
        user_id = user.user.id
        user_email = user.user.email

        # ② 보안을 위해 현재 비밀번호로 재인증 (본인 확인)
        try:
            supabase.auth.sign_in_with_password({
                "email": user_email,
                "password": request.currentPassword
            })
        except Exception:
            raise HTTPException(status_code=400, detail="현재 비밀번호가 일치하지 않습니다.")

        # ③ 비밀번호 업데이트 (Admin 권한 사용)
        # SUPABASE_SERVICE_ROLE_KEY로 생성된 클라이언트여야 합니다.
        supabase.auth.admin.update_user_by_id(
            user_id, 
            {"password": request.newPassword}
        )
        
        return {"message": "성공"}

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

from datetime import date

@app.get("/schedules")
async def get_schedules():
    # 1. DB에서 전체 일정 조회
    response = supabase.table("tax_schedules").select("*").order("due_date").execute()
    schedules = response.data
    
    today = date.today()
    
    # 2. 각 일정별 D-Day 계산 로직 추가
    for item in schedules:
        due = date.fromisoformat(item['due_date'])
        delta = (due - today).days
        item['d_day'] = delta  # 0이면 당일, 양수면 남은 날짜, 음수면 지난 날짜
        
    return schedules

import holidays
from datetime import date, timedelta

# 한국 공휴일 객체 생성
kr_holidays = holidays.KR()

@app.get("/schedules")
async def get_schedules():
    # 1. DB에서 사용자 정의 일정 가져오기
    response = supabase.table("tax_schedules").select("*").order("due_date").execute()
    db_schedules = response.data
    
    today = date.today()
    
    # 2. 이번 달 전후 1년치 공휴일 자동 생성 (캐싱 효과)
    auto_holidays = []
    current_year = today.year
    for date_val, name in holidays.KR(years=[current_year, current_year + 1]).items():
        auto_holidays.append({
            "id": f"holiday-{date_val}",
            "title": name,
            "due_date": date_val.isoformat(),
            "category": "공휴일",
            "is_important": True,
            "target_entity": "대한민국",
            "description": "법정공휴일"
        })

    # 3. DB 일정 + 자동 공휴일 합치기
    all_schedules = db_schedules + auto_holidays
    
    # 4. D-Day 계산
    for item in all_schedules:
        due = date.fromisoformat(item['due_date'])
        item['d_day'] = (due - today).days
        
    return all_schedules
